#!/usr/bin/env python3
import argparse
from collections import defaultdict
import os
import re

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
except ImportError:
    raise SystemExit("The 'openpyxl' library is required. Install it with: pip install openpyxl")


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_RESULTS_DIR = os.path.normpath(
    os.path.join(SCRIPT_DIR, "..", "results", "sms_ai_ml", "with-sms")
)
DEFAULT_OUTPUT_FILE = "l1d_prefetcher_accuracy.xlsx"


def natural_sort_key(value):
    return [int(text) if text.isdigit() else text.lower()
            for text in re.split(r"([0-9]+)", value)]


def parse_file(filepath):
    """
    Parses a single ChampSim output file to find the L1D prefetcher accuracy stats.
    Example line to match:
    L1D USEFUL LOAD PREFETCHES:    6733170 PREFETCH ISSUED TO LOWER LEVEL:    7342702  ACCURACY: 91.6988
    """
    with open(filepath, "r", errors="ignore") as file:
        content = file.read()

    # Regex to match the prefetcher stats
    pattern = (
        r"L1D\s*USEFUL\s*LOAD\s*PREFETCHES:\s*(\d+)\s*"
        r"PREFETCH\s*ISSUED\s*TO\s*LOWER\s*LEVEL:\s*(\d+)\s*"
        r"ACCURACY:\s*([^\s\n\r]+)"
    )
    match = re.search(pattern, content)
    if match:
        useful = int(match.group(1))
        issued = int(match.group(2))
        accuracy_str = match.group(3)
        try:
            accuracy = float(accuracy_str)
        except ValueError:
            accuracy = accuracy_str  # Keep string like "-nan"
        return useful, issued, accuracy
    return None


def collect_records(results_dir):
    records = []
    for root, _, files in os.walk(results_dir):
        for filename in sorted(files, key=natural_sort_key):
            filepath = os.path.join(root, filename)
            if not os.path.isfile(filepath):
                continue

            stats = parse_file(filepath)
            if stats is None:
                continue

            useful, issued, accuracy = stats
            rel_dir = os.path.relpath(root, results_dir)
            trace = filename if rel_dir == "." else os.path.join(rel_dir, filename)
            directory_parts = [] if rel_dir == "." else rel_dir.split(os.sep)
            
            configuration_depth = 2
            if len(directory_parts) > 2 and directory_parts[1] == "pref_l1_l2":
                configuration_depth = 3
                
            configuration = (
                os.path.join(*directory_parts[:configuration_depth])
                if directory_parts
                else "."
            )
            
            records.append({
                "trace": trace,
                "trace_name": filename,
                "configuration": configuration,
                "path": filepath,
                "useful": useful,
                "issued": issued,
                "accuracy": accuracy
            })
    return records


def safe_sheet_name(configuration, used_names):
    name = re.sub(r"[\\/*?:\[\]]", "_", configuration).strip("_") or "results"
    name = name[:31]
    candidate = name
    suffix = 2
    while candidate.lower() in used_names:
        suffix_text = f"_{suffix}"
        candidate = f"{name[:31 - len(suffix_text)]}{suffix_text}"
        suffix += 1
    used_names.add(candidate.lower())
    return candidate


def write_workbook(records, output_file):
    grouped_records = defaultdict(list)
    for record in records:
        grouped_records[record["configuration"]].append(record)

    workbook = Workbook()
    workbook.remove(workbook.active)
    used_names = set()

    for configuration in sorted(grouped_records, key=natural_sort_key):
        worksheet = workbook.create_sheet(safe_sheet_name(configuration, used_names))
        worksheet.cell(1, 1, configuration.replace(os.sep, "/"))
        worksheet.cell(1, 1).font = Font(bold=True)

        headers_row = 3
        headers = [
            "Trace",
            "L1D Useful Load Prefetches",
            "Prefetch Issued to Lower Level",
            "Accuracy (%)"
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(headers_row, col_idx, header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        row_idx = 4
        for record in sorted(grouped_records[configuration], key=lambda item: natural_sort_key(item["trace_name"])):
            worksheet.cell(row_idx, 1, record["trace_name"])
            worksheet.cell(row_idx, 2, record["useful"])
            worksheet.cell(row_idx, 3, record["issued"])
            
            acc_cell = worksheet.cell(row_idx, 4, record["accuracy"])
            if isinstance(record["accuracy"], float):
                acc_cell.number_format = "0.0000"
            else:
                acc_cell.alignment = Alignment(horizontal="right")
            row_idx += 1

        worksheet.freeze_panes = "B4"
        worksheet.column_dimensions["A"].width = 48
        worksheet.column_dimensions["B"].width = 28
        worksheet.column_dimensions["C"].width = 32
        worksheet.column_dimensions["D"].width = 16

    workbook.save(output_file)


def main():
    parser = argparse.ArgumentParser(
        description="Extract L1D prefetcher accuracy statistics from ChampSim output files."
    )
    parser.add_argument(
        "results_dir",
        nargs="?",
        default=DEFAULT_RESULTS_DIR,
        help=f"Directory containing ChampSim output files. Default: {DEFAULT_RESULTS_DIR}",
    )
    parser.add_argument(
        "-o",
        "--output",
        default=DEFAULT_OUTPUT_FILE,
        help=f"Excel output file. Default: {DEFAULT_OUTPUT_FILE}",
    )
    args = parser.parse_args()

    if not os.path.isdir(args.results_dir):
        raise SystemExit(f"Results directory not found: {args.results_dir}")

    records = collect_records(args.results_dir)
    if not records:
        raise SystemExit(
            "No parseable result files found under: "
            f"{args.results_dir}\n"
            "Expected lines containing 'L1D USEFUL LOAD PREFETCHES:'."
        )

    # Resolve output path relative to script directory if not absolute
    output_path = args.output
    if not os.path.isabs(output_path):
        output_path = os.path.join(SCRIPT_DIR, output_path)

    write_workbook(records, output_path)
    print(f"Successfully processed {len(records)} trace files.")
    print(f"Data saved to Excel spreadsheet: {output_path}")


if __name__ == "__main__":
    main()
