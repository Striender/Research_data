#!/usr/bin/env python3
import argparse
from collections import defaultdict
import os
import re

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
except ImportError:
    raise SystemExit(
        "The 'openpyxl' library is required. Install it with: pip install openpyxl"
    )


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_RESULTS_DIR = os.path.normpath(
    os.path.join(SCRIPT_DIR, "..", "results", "speedup", "test")
)
DEFAULT_OUTPUT_FILE = "../Excel_Output/Pref_l2/test_Prefetcher_accuracy.xlsx"
CACHE_LEVELS = ("L1D", "L2C", "LLC")


def natural_sort_key(value):
    return [
        int(text) if text.isdigit() else text.lower()
        for text in re.split(r"([0-9]+)", value)
    ]


def parse_file(filepath):
    """
    Parses a single ChampSim output file to extract prefetch statistics for
    L1D, L2C, and LLC.

    Expected lines:

    L1D PREFETCH  REQUESTED:    1189392  ISSUED:    1189392  USEFUL:     813535  USELESS:       1128
    L1D USEFUL LOAD PREFETCHES:     813535 PREFETCH ISSUED TO LOWER LEVEL:    1117280  ACCURACY: 72.8139
    L1D TIMELY PREFETCHES:     813535 LATE PREFETCHES: 302624 DROPPED PREFETCHES: 0
    """

    with open(filepath, "r", errors="ignore") as file:
        content = file.read()

    metrics = {}

    for cache_name in CACHE_LEVELS:
        accuracy_pattern = (
            rf"{cache_name}\s*USEFUL\s*LOAD\s*PREFETCHES:\s*(\d+)\s*"
            r"PREFETCH\s*ISSUED\s*TO\s*LOWER\s*LEVEL:\s*(\d+)\s*"
            r"ACCURACY:\s*([^\s\n\r]+)"
        )

        prefetch_pattern = (
            rf"{cache_name}\s*PREFETCH\s*REQUESTED:\s*\d+\s*"
            r"ISSUED:\s*\d+\s*"
            r"USEFUL:\s*(\d+)\s*"
            r"USELESS:\s*(\d+)"
            r"(?:\s*total_filled\s+by\s+prefetcher:\s*(\d+))?"
        )

        late_pattern = (
            rf"{cache_name}\s*TIMELY\s*PREFETCHES:\s*\d+\s*"
            r"LATE\s*PREFETCHES:\s*(\d+)"
        )

        acc_match = re.search(accuracy_pattern, content)
        pref_match = re.search(prefetch_pattern, content)
        late_match = re.search(late_pattern, content)

        if not (acc_match and pref_match and late_match):
            continue

        accuracy_str = acc_match.group(3)
        try:
            accuracy = float(accuracy_str)
        except ValueError:
            accuracy = accuracy_str

        metrics[cache_name] = {
            "useful": int(acc_match.group(1)),
            "issued": int(acc_match.group(2)),
            "accuracy": accuracy,
            "useless": int(pref_match.group(2)),
            "total_filled": int(pref_match.group(3)) if pref_match.group(3) else "N/A",
            "late": int(late_match.group(1)),
        }

    return metrics or None


def collect_records(results_dir):
    records = []

    for root, _, files in os.walk(results_dir):
        for filename in sorted(files, key=natural_sort_key):
            filepath = os.path.join(root, filename)

            if not os.path.isfile(filepath):
                continue

            stats = parse_file(filepath)
            if stats is None:
                continue

            rel_dir = os.path.relpath(root, results_dir)
            trace = filename if rel_dir == "." else os.path.join(rel_dir, filename)

            directory_parts = [] if rel_dir == "." else rel_dir.split(os.sep)

            configuration_depth = 2
            if len(directory_parts) > 2 and directory_parts[1] == "pref_l1_l2":
                configuration_depth = 3

            configuration = (
                os.path.join(*directory_parts[:configuration_depth])
                if directory_parts
                else "."
            )

            records.append(
                {
                    "trace": trace,
                    "trace_name": filename,
                    "configuration": configuration,
                    "path": filepath,
                    "metrics": stats,
                }
            )

    return records


def safe_sheet_name(configuration, used_names):
    name = re.sub(r"[\\/*?:\[\]]", "_", configuration).strip("_") or "results"
    name = name[:31]

    candidate = name
    suffix = 2

    while candidate.lower() in used_names:
        suffix_text = f"_{suffix}"
        candidate = f"{name[:31-len(suffix_text)]}{suffix_text}"
        suffix += 1

    used_names.add(candidate.lower())
    return candidate


def write_workbook(records, output_file):
    grouped_records = defaultdict(list)

    for record in records:
        grouped_records[record["configuration"]].append(record)

    workbook = Workbook()
    workbook.remove(workbook.active)

    used_names = set()

    for configuration in sorted(grouped_records, key=natural_sort_key):
        worksheet = workbook.create_sheet(
            safe_sheet_name(configuration, used_names)
        )

        worksheet.cell(1, 1, configuration.replace(os.sep, "/"))
        worksheet.cell(1, 1).font = Font(bold=True)

        headers_row = 3

        headers = [
            "Trace",
        ]

        for cache_name in CACHE_LEVELS:
            headers.extend(
                [
                    f"{cache_name} Useful Load Prefetches",
                    f"{cache_name} Prefetch Issued to Lower Level",
                    f"{cache_name} Accuracy (%)",
                    f"{cache_name} Useless Prefetches",
                    f"{cache_name} Total Filled by Prefetcher",
                    f"{cache_name} Late Prefetches",
                ]
            )

        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(headers_row, col_idx, header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        row_idx = 4

        for record in sorted(
            grouped_records[configuration],
            key=lambda item: natural_sort_key(item["trace_name"]),
        ):
            worksheet.cell(row_idx, 1, record["trace_name"])

            col_idx = 2
            for cache_name in CACHE_LEVELS:
                stats = record["metrics"].get(cache_name)
                if stats is None:
                    for offset in range(6):
                        worksheet.cell(row_idx, col_idx + offset, "N/A")
                    col_idx += 6
                    continue

                worksheet.cell(row_idx, col_idx, stats["useful"])
                worksheet.cell(row_idx, col_idx + 1, stats["issued"])

                acc_cell = worksheet.cell(row_idx, col_idx + 2, stats["accuracy"])
                if isinstance(stats["accuracy"], float):
                    acc_cell.number_format = "0.0000"
                else:
                    acc_cell.alignment = Alignment(horizontal="right")

                worksheet.cell(row_idx, col_idx + 3, stats["useless"])
                worksheet.cell(row_idx, col_idx + 4, stats["total_filled"])
                worksheet.cell(row_idx, col_idx + 5, stats["late"])
                col_idx += 6

            row_idx += 1

        worksheet.freeze_panes = "B4"

        worksheet.column_dimensions["A"].width = 48

        for column_cells in worksheet.columns:
            column_letter = column_cells[0].column_letter
            if column_letter == "A":
                continue
            max_len = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            worksheet.column_dimensions[column_letter].width = min(max_len + 2, 34)

    workbook.save(output_file)


def main():
    parser = argparse.ArgumentParser(
        description="Extract L1D, L2C, and LLC prefetch statistics from ChampSim output files."
    )

    parser.add_argument(
        "results_dir",
        nargs="?",
        default=DEFAULT_RESULTS_DIR,
        help=f"Directory containing ChampSim output files. Default: {DEFAULT_RESULTS_DIR}",
    )

    parser.add_argument(
        "-o",
        "--output",
        default=DEFAULT_OUTPUT_FILE,
        help=f"Excel output file. Default: {DEFAULT_OUTPUT_FILE}",
    )

    args = parser.parse_args()

    if not os.path.isdir(args.results_dir):
        raise SystemExit(f"Results directory not found: {args.results_dir}")

    records = collect_records(args.results_dir)

    if not records:
        raise SystemExit(
            "No parseable result files found under: "
            f"{args.results_dir}\n"
            "Expected lines containing '<cache> USEFUL LOAD PREFETCHES:'."
        )

    output_path = args.output

    if not os.path.isabs(output_path):
        output_path = os.path.join(SCRIPT_DIR, output_path)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    write_workbook(records, output_path)

    print(f"Successfully processed {len(records)} trace files.")
    print(f"Data saved to Excel spreadsheet: {output_path}")


if __name__ == "__main__":
    main()
