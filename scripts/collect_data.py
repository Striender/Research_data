#!/usr/bin/env python3
import os
import re
import pandas as pd
from collections import defaultdict
import json

# Try to import openpyxl and guide the user if it's not installed.
try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.cell import MergedCell
except ImportError:
    print("The 'openpyxl' library is required to write formatted Excel files.")
    print("Please install it on your server by running: pip install openpyxl")
    exit()

# ANSI escape codes for terminal colors
class TColors:
    OKGREEN = '\033[92m'
    ENDC = '\033[0m'

def natural_sort_key(s):
    """
    Create a sort key that handles numbers inside strings for natural sorting.
    e.g., 'exp1', 'exp2', 'exp10'
    """
    return [int(text) if text.isdigit() else text.lower() for text in re.split('([0-9]+)', s)]

def load_json_data(file_path):
    """Loads a JSON file."""
    if not os.path.exists(file_path):
        return {}
    try:
        with open(file_path, 'r') as f:
            return json.load(f)
    except (IOError, json.JSONDecodeError):
        # Return empty dict if file is corrupted or empty
        return {}

def save_json_data(file_path, data):
    """Saves data to a JSON file."""
    try:
        with open(file_path, 'w') as f:
            json.dump(data, f, indent=4)
    except IOError as e:
        print(f"Warning: Could not save the cache/log file at {file_path}: {e}")

def parse_champsim_file(filepath):
    """
    Parses a single ChampSim output file to extract specific metrics.

    Args:
        filepath (str): The full path to the ChampSim output file.

    Returns:
        dict: A dictionary containing the extracted metrics.
    """
    # Initialize a dictionary with default values for all metrics
    metrics = {
        "Trace File": os.path.basename(filepath),
        "IPC": None,
        "LLC Total Access": None,
        "LLC Total Hits": None,
        "LLC Total Misses": None,
        "LLC Total MPKI": None,
        "LLC Load Access": None,
        "LLC Load Hit": None,
        "LLC Load Miss": None,
        "LLC Load MPKI": None,
        "L2C Data Load MPKI": None,
        "L1D Prefetch Requested": None,
        "L1D Prefetch Issued": None,
        "L1D Prefetch Useful": None,
        "L1D Prefetch Useless": None,
        "L1D Useful Load Prefetches": None,
        "L1D Timely Prefetches": None,
        "L1D Late Prefetches": None,
        "L1D Dropped Prefetches": None,
        "LLC Average Miss Latency": None
    }

    try:
        with open(filepath, 'r', errors='ignore') as f:
            content = f.read()

            # --- Extract IPC ---
            ipc_match = re.search(r"CPU 0 cumulative IPC:\s+([\d.]+)", content)
            if ipc_match:
                metrics["IPC"] = float(ipc_match.group(1))

            # --- LLC Total Stats ---
            llc_total_match = re.search(
                r"LLC TOTAL\s+ACCESS:\s+(\d+)\s+HIT:\s+(\d+)\s+MISS:\s+(\d+).*?MPKI:\s+([\d.]+)",
                content
            )
            if llc_total_match:
                metrics["LLC Total Access"] = int(llc_total_match.group(1))
                metrics["LLC Total Hits"] = int(llc_total_match.group(2))
                metrics["LLC Total Misses"] = int(llc_total_match.group(3))
                metrics["LLC Total MPKI"] = float(llc_total_match.group(4))

            # --- LLC Load Stats ---
            llc_load_match = re.search(
                r"LLC LOAD\s+ACCESS:\s+(\d+)\s+HIT:\s+(\d+)\s+MISS:\s+(\d+).*?MPKI:\s+([\d.]+)",
                content
            )
            if llc_load_match:
                metrics["LLC Load Access"] = int(llc_load_match.group(1))
                metrics["LLC Load Hit"] = int(llc_load_match.group(2))
                metrics["LLC Load Miss"] = int(llc_load_match.group(3))
                metrics["LLC Load MPKI"] = float(llc_load_match.group(4))
            
            # --- L2C Data Load MPKI ---
            l2c_data_load_mpki_match = re.search(r"L2C DATA LOAD MPKI:\s+([\d.]+)", content)
            if l2c_data_load_mpki_match:
                metrics["L2C Data Load MPKI"] = float(l2c_data_load_mpki_match.group(1))

            # --- L1D Prefetch Requested/Issued/Useful/Useless ---
            l1d_prefetch_match = re.search(
                r"L1D PREFETCH\s+REQUESTED:\s+(\d+)\s+ISSUED:\s+(\d+)\s+USEFUL:\s+(\d+)\s+USELESS:\s+(\d+)",
                content
            )
            if l1d_prefetch_match:
                metrics["L1D Prefetch Requested"] = int(l1d_prefetch_match.group(1))
                metrics["L1D Prefetch Issued"] = int(l1d_prefetch_match.group(2))
                metrics["L1D Prefetch Useful"] = int(l1d_prefetch_match.group(3))
                metrics["L1D Prefetch Useless"] = int(l1d_prefetch_match.group(4))

            # --- L1D Useful/Timely/Late/Dropped Prefetches ---
            l1d_useful_load_match = re.search(r"L1D USEFUL LOAD PREFETCHES:\s+(\d+)", content)
            if l1d_useful_load_match:
                metrics["L1D Useful Load Prefetches"] = int(l1d_useful_load_match.group(1))

            l1d_timely_match = re.search(
                r"L1D TIMELY PREFETCHES:\s+(\d+)\s+LATE PREFETCHES:\s+(\d+)\s+DROPPED PREFETCHES:\s+(\d+)",
                content
            )
            if l1d_timely_match:
                metrics["L1D Timely Prefetches"] = int(l1d_timely_match.group(1))
                metrics["L1D Late Prefetches"] = int(l1d_timely_match.group(2))
                metrics["L1D Dropped Prefetches"] = int(l1d_timely_match.group(3))

            # --- LLC Average Miss Latency ---
            llc_latency_match = re.search(r"LLC AVERAGE MISS LATENCY:\s+([\d.]+)", content)
            if llc_latency_match:
                metrics["LLC Average Miss Latency"] = float(llc_latency_match.group(1))

    except IOError as e:
        print(f"Error reading file {filepath}: {e}")
        return None
        
    return metrics

def apply_border_to_range(worksheet, row_range, col_range, border_style):
    """Helper function to apply a border to a range of cells."""
    for row in worksheet.iter_rows(min_row=row_range[0], max_row=row_range[1], min_col=col_range[0], max_col=col_range[1]):
        for cell in row:
            cell.border = border_style

def main():
    """
    Main function to find ChampSim files, parse them, and save them to a single
    formatted Excel file with multiple sheets, skipping already processed files.
    """
    # --- CONFIGURATION ---
    RESULTS_DIR = "../results/"
    OUTPUT_DIR = "../Extracted_Data/"
    EXCEL_OUTPUT_FILE = "rnd.xlsx"
    PROCESSED_LOG_FILE = os.path.join(OUTPUT_DIR, ".processed_files.log")
    DATA_CACHE_FILE = os.path.join(OUTPUT_DIR, ".data_cache.json")
    # -------------------

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    if not os.path.isdir(RESULTS_DIR):
        print(f"Error: Directory '{RESULTS_DIR}' not found.")
        return

    processed_files_log = load_json_data(PROCESSED_LOG_FILE)
    cached_data = load_json_data(DATA_CACHE_FILE)
    
    # The structure is {group_key: {filepath: metrics}}
    # We will rebuild this dict completely to handle file deletions properly
    data_by_prefetcher = defaultdict(dict)
    
    new_files_count = 0
    skipped_files_count = 0
    announced_dirs = set() # To prevent printing the same directory multiple times
    
    print(f"Starting scan in directory: '{RESULTS_DIR}'...")
    # Walk through the directory tree to collect all data
    for root, dirs, files in os.walk(RESULTS_DIR):
        if not files: continue

        relative_path = os.path.relpath(root, RESULTS_DIR)
        path_parts = relative_path.split(os.sep)

        group_key, experiment = None, None

        if len(path_parts) == 3: # Standard case: results/pref_l1/berti/exp1
            cache_level, prefetcher, experiment = path_parts
            group_key = f"{cache_level}_{prefetcher}"
        elif len(path_parts) == 2: # Edge case for no_pref: results/no_pref/exp1
            cache_level, experiment = path_parts
            if cache_level == 'no_pref': group_key = cache_level
        
        if group_key and experiment:
            for filename in sorted(files):
                filepath = os.path.join(root, filename)
                
                file_mod_time = os.path.getmtime(filepath)
                # If file is unchanged, load its data from cache instead of re-parsing
                if processed_files_log.get(filepath) == file_mod_time:
                    if group_key in cached_data and filepath in cached_data[group_key]:
                        data_by_prefetcher[group_key][filepath] = cached_data[group_key][filepath]
                    skipped_files_count += 1
                    continue
                
                # If file is new or modified, announce the directory once
                if root not in announced_dirs:
                    print(f"{TColors.OKGREEN}Processing new/modified files in: {relative_path}{TColors.ENDC}")
                    announced_dirs.add(root)

                new_files_count += 1
                metrics = parse_champsim_file(filepath)
                if metrics:
                    metrics["Experiment"] = experiment
                    data_by_prefetcher[group_key][filepath] = metrics
                    processed_files_log[filepath] = file_mod_time

    print(f"\nScan complete. Found {new_files_count} new/modified files. Skipped {skipped_files_count} unchanged files.")

    if new_files_count == 0:
        print("\nOutput is already up-to-date.")
        return

    print(f"\nProcessing data and updating {EXCEL_OUTPUT_FILE}...")
    output_path = os.path.join(OUTPUT_DIR, EXCEL_OUTPUT_FILE)
    
    try:
        # Load existing workbook or create a new one
        if os.path.exists(output_path):
            book = load_workbook(output_path)
            # Create a list of sheet names to manage them, as we'll be modifying the book
            sheet_names = book.sheetnames
        else:
            from openpyxl import Workbook
            book = Workbook()
            sheet_names = []
            if 'Sheet' in book.sheetnames:
                 book.remove(book.active)


        # Define styles once
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        main_header_fill = PatternFill(start_color="FFDAB9", end_color="FFDAB9", fill_type="solid") # Peach
        data_header_fill = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid") # Dark Gray
        sub_header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid") # Light Blue
        center_alignment = Alignment(horizontal='center', vertical='center')

        for group_key in sorted(data_by_prefetcher.keys()):
            data_list = list(data_by_prefetcher[group_key].values())
            if not data_list: continue

            print(f"Processing group: {group_key}")
            
            df = pd.DataFrame(data_list)
            sheet_name = f"raw_{group_key}"

            # If sheet exists, remove it to be rewritten
            if sheet_name in sheet_names:
                book.remove(book[sheet_name])
            
            # Create a new sheet for the data
            worksheet = book.create_sheet(title=sheet_name)

            df_for_headers = df.drop(['Experiment'], axis=1, errors='ignore')
            
            # --- Create and Write Main Header ---
            if group_key == 'no_pref':
                main_header_text = "Baseline (No Prefetcher)"
            else:
                parts = group_key.split('_')
                cache_level_str = parts[1].upper() if len(parts) > 1 else ''
                prefetcher_name = '_'.join(parts[2:]).capitalize() if len(parts) > 2 else parts[0].capitalize()
                main_header_text = f"Data Prefetcher: {prefetcher_name} at {cache_level_str}"
            
            num_cols = len(df_for_headers.columns)

            worksheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=num_cols)
            main_header_cell = worksheet.cell(row=1, column=1, value=main_header_text)
            main_header_cell.font = Font(bold=True, size=14)
            main_header_cell.alignment = center_alignment
            main_header_cell.fill = main_header_fill
            apply_border_to_range(worksheet, (1, 2), (1, num_cols), thin_border)

            # Write and Style the data column headers on row 3
            for col_num, col_name in enumerate(df_for_headers.columns, 1):
                cell = worksheet.cell(row=3, column=col_num, value=col_name)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = data_header_fill
                cell.border = thin_border
                cell.alignment = center_alignment
            worksheet.row_dimensions[3].height = 30
            
            worksheet.freeze_panes = 'A4' # Freeze rows 1, 2, and 3

            experiments = sorted(df['Experiment'].unique(), key=natural_sort_key)
            current_row = 3

            for experiment in experiments:
                df_experiment = df[df['Experiment'] == experiment].copy()
                
                if current_row > 3: current_row += 1 

                bold_font = Font(bold=True, size=12)
                
                # --- Create Descriptive Experiment Header ---
                try:
                    exp_parts = experiment.split('_')
                    exp_num = ''.join(filter(str.isdigit, exp_parts[0]))
                    l2_policy = exp_parts[1].upper()
                    llc_policy = exp_parts[2].upper()
                    exp_header_text = f"Experiment {exp_num}: Replacement Policy {l2_policy} at L2 and {llc_policy} at LLC"
                except (IndexError, ValueError):
                    exp_header_text = experiment.replace('_', ' ').title()

                worksheet.merge_cells(start_row=current_row + 1, start_column=1, end_row=current_row + 1, end_column=num_cols)
                
                header_cell = worksheet.cell(row=current_row + 1, column=1, value=exp_header_text)
                header_cell.font = bold_font
                header_cell.alignment = center_alignment
                header_cell.fill = sub_header_fill
                apply_border_to_range(worksheet, (current_row + 1, current_row + 1), (1, num_cols), thin_border)
                worksheet.row_dimensions[current_row + 1].height = 30

                df_to_write = df_experiment.drop(['Experiment'], axis=1, errors='ignore')
                
                # Write data using openpyxl to avoid pandas overwriting styles
                for r_idx, row_data in enumerate(df_to_write.itertuples(index=False), start=current_row + 2):
                    for c_idx, value in enumerate(row_data, 1):
                        worksheet.cell(row=r_idx, column=c_idx, value=value)
                
                current_row += 1 + len(df_experiment)
            
            bold_font_for_trace = Font(bold=True)
            for cell in worksheet['A']:
                if cell.row > 3 and cell.value and worksheet.cell(row=cell.row, column=2).value:
                    cell.font = bold_font_for_trace

            for col_idx in range(1, worksheet.max_column + 1):
                column_letter = get_column_letter(col_idx)
                max_length = 0
                for cell in worksheet[column_letter]:
                    if isinstance(cell, MergedCell): continue
                    if cell.value: max_length = max(len(str(cell.value)), max_length)
                worksheet.column_dimensions[column_letter].width = max_length + 2
            
            print(f" -> Finished writing and formatting sheet: {sheet_name}")
        
        # Save the entire workbook at the end
        book.save(output_path)

        # Clean the cache for saving by removing internal flags
        for group in data_by_prefetcher.values():
            for record in group.values():
                record.pop('_is_new', None)
                
        save_json_data(DATA_CACHE_FILE, data_by_prefetcher)
        save_json_data(PROCESSED_LOG_FILE, processed_files_log)
        print(f"\nSuccessfully created/updated Excel file: {output_path}")

    except Exception as e:
        print(f"\nAn error occurred while writing the Excel file: {e}")

if __name__ == "__main__":
    main()
