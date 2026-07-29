#!/usr/bin/env python3
"""Bin SMS PHT insertion frequencies into 1 and >1 categories."""

from __future__ import annotations

import argparse
import os
import re
import shutil
import tempfile
from collections import defaultdict
from pathlib import Path


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_RESULTS_DIR = os.path.normpath(
    os.path.join(SCRIPT_DIR, "..", "results", "SMS_region", "2k")
)
DEFAULT_OUTPUT_FILE = os.path.normpath(
    os.path.join(SCRIPT_DIR, "..", "Excel_Output", "sms", "PHT_2k_insert_frequency.xlsx")
)

KEY_FREQUENCY_RE = re.compile(
    r"^\s*sms\.pht\.key\s+\S+\s+insert_frequency\s+(\d+)\s*$", re.MULTILINE
)
PHT_INSERT_RE = re.compile(r"^\s*sms\.pht\.insert\s+(\d+)\s*$", re.MULTILINE)
NUMERIC_BLOCK_RE = re.compile(
    r"^\s*sms\.pref_buffer\.issued\s+\d+\s*$\n(.*?)(?:=){6,}",
    re.MULTILINE | re.DOTALL,
)

HEADERS = [
    "Trace File", "Bin 1 Frequency", "Bin 1 Percentage",
    "Bin >1 Frequency", "Bin >1 Percentage",
]


def natural_sort_key(value: str) -> list[object]:
    return [
        int(part) if part.isdigit() else part.casefold()
        for part in re.split(r"(\d+)", value)
    ]


def numeric_frequency_block(content: str) -> list[int]:
    """Read frequencies after issued, including an inline ===== terminator."""
    candidates: list[list[int]] = []
    for match in NUMERIC_BLOCK_RE.finditer(content):
        block = match.group(1).strip()
        if block and re.fullmatch(r"\d+(?:\s+\d+)*", block):
            candidates.append([int(value) for value in block.split()])
    return candidates[-1] if candidates else []


def parse_trace(trace_path: Path) -> dict[str, object] | None:
    try:
        content = trace_path.read_text(errors="ignore")
    except OSError as exc:
        print(f"Warning: could not read {trace_path}: {exc}")
        return None

    keyed_values = [int(value) for value in KEY_FREQUENCY_RE.findall(content)]
    if keyed_values:
        frequencies = keyed_values
        frequency_format = "keyed"
    else:
        frequencies = numeric_frequency_block(content)
        frequency_format = "numeric block"
    if not frequencies:
        return None

    insert_matches = PHT_INSERT_RE.findall(content)
    pht_insert = int(insert_matches[-1]) if insert_matches else None
    frequency_sum = sum(frequencies)
    denominator = pht_insert if pht_insert is not None else frequency_sum
    denominator_source = "sms.pht.insert" if pht_insert is not None else "frequency sum"

    bin_one = sum(value for value in frequencies if value == 1)
    bin_greater = sum(value for value in frequencies if value > 1)
    # Store percentages as fractions because Excel's percentage number format
    # performs the x100 conversion when displaying the cell.
    percent_one = bin_one / denominator if denominator else 0.0
    percent_greater = bin_greater / denominator if denominator else 0.0

    return {
        "Frequency Format": frequency_format,
        "Signatures": len(frequencies),
        "PHT Insert": pht_insert,
        "Frequency Sum": frequency_sum,
        "Denominator": denominator,
        "Denominator Source": denominator_source,
        "Bin 1 Frequency": bin_one,
        "Bin 1 Percentage": percent_one,
        "Bin >1 Frequency": bin_greater,
        "Bin >1 Percentage": percent_greater,
        "Totals Match": pht_insert is None or pht_insert == frequency_sum,
    }


def collect_results(results_dir: Path) -> tuple[dict[str, list[dict[str, object]]], int]:
    grouped: dict[str, list[dict[str, object]]] = defaultdict(list)
    scanned = 0
    for root, _dirs, files in os.walk(results_dir):
        relative_path = os.path.relpath(root, results_dir)
        path_parts = relative_path.split(os.sep)
        group_key = experiment = None
        if len(path_parts) == 4:
            cache, size, prefetcher, experiment = path_parts
            group_key = f"{cache}_{size}_{prefetcher}"
        elif len(path_parts) == 3:
            cache_level, prefetcher, experiment = path_parts
            group_key = f"{cache_level}_{prefetcher}"
        elif len(path_parts) == 2:
            cache_level, experiment = path_parts
            if cache_level == "no_pref":
                group_key = cache_level
        if not group_key or not experiment:
            continue

        for filename in sorted(files, key=natural_sort_key):
            scanned += 1
            metrics = parse_trace(Path(root) / filename)
            if metrics:
                metrics["Trace File"] = filename
                metrics["Experiment"] = experiment
                grouped[group_key].append(metrics)
    return dict(grouped), scanned


def experiment_title(experiment: str) -> str:
    parts = experiment.split("_")
    if len(parts) >= 3:
        number = "".join(filter(str.isdigit, parts[0]))
        return (
            f"Experiment {number}: Replacement Policy {parts[1].upper()} at L2 "
            f"and {parts[2].upper()} at LLC"
        )
    return experiment.replace("_", " ").title()


def write_workbook(grouped: dict[str, list[dict[str, object]]], output: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.cell import MergedCell
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("Install openpyxl to write the workbook: pip install openpyxl") from exc

    output.parent.mkdir(parents=True, exist_ok=True)
    book = Workbook()
    book.active.title = "Placeholder"
    border = Border(**{side: Side(style="thin") for side in ("left", "right", "top", "bottom")})
    peach = PatternFill("solid", fgColor="FFDAB9")
    gray = PatternFill("solid", fgColor="A9A9A9")
    blue = PatternFill("solid", fgColor="ADD8E6")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    def border_range(sheet, row1, row2, col1, col2):
        for row in sheet.iter_rows(min_row=row1, max_row=row2, min_col=col1, max_col=col2):
            for cell in row:
                cell.border = border

    for group_key in sorted(grouped, key=natural_sort_key):
        sheet = book.create_sheet(group_key[:31])
        sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=len(HEADERS))
        title = sheet.cell(1, 1, f"SMS PHT Insert Frequency: {group_key}")
        title.font = Font(bold=True, size=14)
        title.fill = peach
        title.alignment = center
        border_range(sheet, 1, 2, 1, len(HEADERS))
        for column, header in enumerate(HEADERS, 1):
            cell = sheet.cell(3, column, header)
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.fill = gray
            cell.border = border
            cell.alignment = center
        sheet.freeze_panes = "A4"

        row_number = 4
        experiments = sorted({str(r["Experiment"]) for r in grouped[group_key]}, key=natural_sort_key)
        for experiment in experiments:
            sheet.merge_cells(
                start_row=row_number, start_column=1,
                end_row=row_number, end_column=len(HEADERS),
            )
            cell = sheet.cell(row_number, 1, experiment_title(experiment))
            cell.font = Font(bold=True, size=12)
            cell.fill = blue
            cell.alignment = center
            border_range(sheet, row_number, row_number, 1, len(HEADERS))
            row_number += 1
            records = [r for r in grouped[group_key] if r["Experiment"] == experiment]
            records.sort(key=lambda r: natural_sort_key(str(r["Trace File"])))
            for record in records:
                for column, header in enumerate(HEADERS, 1):
                    cell = sheet.cell(row_number, column, record.get(header))
                    cell.alignment = left if column == 1 else right
                    if column == 1:
                        cell.font = Font(bold=True)
                    if header in ("Bin 1 Percentage", "Bin >1 Percentage"):
                        cell.number_format = "0.00%"
                row_number += 1
            row_number += 1

        for column in range(1, len(HEADERS) + 1):
            letter = get_column_letter(column)
            lengths = [
                len(str(cell.value)) for cell in sheet[letter]
                if not isinstance(cell, MergedCell) and cell.value is not None
            ]
            sheet.column_dimensions[letter].width = min(max(lengths, default=10) + 2, 55)

    summary_headers = ["Experiment", "Trace File", "Metric"] + sorted(grouped, key=natural_sort_key)
    summary = book.create_sheet("Frequency_Summary", 0)
    for column, header in enumerate(summary_headers, 1):
        cell = summary.cell(1, column, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = gray
        cell.border = border
        cell.alignment = center
    summary_values: dict[tuple[str, str], dict[str, dict[str, object]]] = defaultdict(dict)
    for group_key, records in grouped.items():
        for record in records:
            summary_values[(str(record["Experiment"]), str(record["Trace File"]))][group_key] = record
    summary_row = 2
    for experiment, trace in sorted(
        summary_values, key=lambda key: (natural_sort_key(key[0]), natural_sort_key(key[1]))
    ):
        for metric in (
            "Bin 1 Frequency", "Bin 1 Percentage",
            "Bin >1 Frequency", "Bin >1 Percentage",
        ):
            values = [experiment, trace, metric]
            values += [summary_values[(experiment, trace)].get(group, {}).get(metric) for group in summary_headers[3:]]
            for column, value in enumerate(values, 1):
                cell = summary.cell(summary_row, column, value)
                cell.border = border
                cell.alignment = left if column <= 3 else right
                if column > 3:
                    cell.number_format = "0.00%"
            summary_row += 1
    summary.freeze_panes = "D2"
    summary.auto_filter.ref = summary.dimensions
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 55
    summary.column_dimensions["C"].width = 22
    for column in range(4, len(summary_headers) + 1):
        summary.column_dimensions[get_column_letter(column)].width = 24

    book.remove(book["Placeholder"])
    temporary = Path(tempfile.gettempdir()) / f"{os.getpid()}_{output.name}"
    book.save(temporary)
    shutil.move(temporary, output)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("results_dir", nargs="?", type=Path, default=Path(DEFAULT_RESULTS_DIR))
    parser.add_argument("-o", "--output", type=Path, default=Path(DEFAULT_OUTPUT_FILE))
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    results_dir = args.results_dir.resolve()
    output = args.output.resolve()
    if not results_dir.is_dir():
        print(f"Error: directory not found: {results_dir}")
        return 1
    grouped, scanned = collect_results(results_dir)
    extracted = sum(len(records) for records in grouped.values())
    if not extracted:
        print("No insertion-frequency data found.")
        return 1
    try:
        write_workbook(grouped, output)
    except RuntimeError as exc:
        print(f"Error: {exc}")
        return 1
    print(f"Extracted {extracted} traces from {scanned} files into {len(grouped)} group sheets.")
    print(f"Workbook written to: {output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
