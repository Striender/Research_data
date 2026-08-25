#!/usr/bin/env python3
import argparse
from collections import defaultdict
import os
import re

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
except ImportError:
    raise SystemExit("The 'openpyxl' library is required. Install it with: pip install openpyxl")


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_RESULTS_DIR = os.path.normpath(
    os.path.join(SCRIPT_DIR, "..", "results", "speedup", "MSHR_FULL_STREAKS")
)
DEFAULT_OUTPUT_FILE = "../Excel_Output/mshr_full_streaks.xlsx"
CACHE_LEVELS = ("L1D", "L2C", "LLC")
SHEET_NAMES = {"L1D": "L1D", "L2C": "L2", "LLC": "LLC"}
STREAK_SHEET_NAMES = {"L1D": "L1D_streaks", "L2C": "L2_streaks", "LLC": "LLC_streaks"}
CYCLES_PER_KILO_INSTRUCTION_DIVISOR = 200_000
STREAK_BINS = (
    ("1-35", 1, 35),
    ("36-70", 36, 70),
    ("71-105", 71, 105),
    ("106-140", 106, 140),
    (">140", 141, None),
)


def natural_sort_key(value):
    return [int(text) if text.isdigit() else text.lower()
            for text in re.split(r"([0-9]+)", value)]


def parse_distribution(bucket_text):
    buckets = {}
    for occupancy, cycles in re.findall(r"(\d+)\s*:\s*(\d+)", bucket_text):
        buckets[int(occupancy)] = int(cycles)
    return buckets


def parse_streak_lengths(streak_text):
    return [int(length) for length in re.findall(r"\d+", streak_text)]


def parse_file(filepath):
    with open(filepath, "r", errors="ignore") as file:
        content = file.read()

    distributions = {}
    streaks = {}
    total_cycles = None

    cycles_match = re.search(
        r"Finished CPU\s+0\s+instructions:\s+\d+\s+cycles:\s+(\d+)",
        content,
    )
    if cycles_match:
        total_cycles = int(cycles_match.group(1))

    for cache_name in CACHE_LEVELS:
        match = re.search(
            rf"{cache_name}\s*MSHR OCCUPANCY DISTRIBUTION:[ \t]*([^\n\r]+)",
            content,
        )
        if match:
            distributions[cache_name] = parse_distribution(match.group(1))

        streak_match = re.search(
            rf"{cache_name}\s*MSHR FULL STREAK LENGTHS:[ \t]*([^\n\r]*)",
            content,
        )
        if streak_match:
            streaks[cache_name] = parse_streak_lengths(streak_match.group(1))

    return distributions, streaks, total_cycles


def collect_records(results_dir):
    records = []
    for root, _, files in os.walk(results_dir):
        for filename in sorted(files, key=natural_sort_key):
            filepath = os.path.join(root, filename)
            if not os.path.isfile(filepath):
                continue

            distributions, streaks, total_cycles = parse_file(filepath)
            if not streaks:
                continue

            rel_dir = os.path.relpath(root, results_dir)
            trace = filename if rel_dir == "." else os.path.join(rel_dir, filename)
            directory_parts = [] if rel_dir == "." else rel_dir.split(os.sep)
            configuration_depth = 2
            if len(directory_parts) > 2 and directory_parts[1] == "pref_l1_l2":
                configuration_depth = 3
            configuration = (
                os.path.join(*directory_parts[:configuration_depth])
                if directory_parts
                else "."
            )
            records.append({
                "trace": trace,
                "trace_name": filename,
                "configuration": configuration,
                "path": filepath,
                "distributions": distributions,
                "streaks": streaks,
                "total_cycles": total_cycles,
            })

    return records


def distribution_percentages(buckets, max_occupancy):
    total_cycles = sum(buckets.values())
    if total_cycles == 0:
        return [0.0] * (max_occupancy + 1)

    return [
        (100.0 * buckets.get(occupancy, 0) / total_cycles)
        for occupancy in range(max_occupancy + 1)
    ]


def streak_bin_counts(streak_lengths):
    counts = []
    for _, start, end in STREAK_BINS:
        if end is None:
            counts.append(sum(1 for length in streak_lengths if length >= start))
        else:
            counts.append(sum(1 for length in streak_lengths if start <= length <= end))
    return counts


def streak_bin_cycles(streak_lengths):
    cycles = []
    for _, start, end in STREAK_BINS:
        if end is None:
            cycles.append(sum(length for length in streak_lengths if length >= start))
        else:
            cycles.append(sum(length for length in streak_lengths if start <= length <= end))
    return cycles


def full_occupancy_cycles(buckets):
    if not buckets:
        return 0
    return buckets[max(buckets)]


def streak_bin_cycle_percentages(cycles, total_full_cycles):
    if total_full_cycles == 0:
        return [0.0] * len(cycles)
    return [(100.0 * cycle_count / total_full_cycles) for cycle_count in cycles]


def mshr_full_cycle_percentage(total_full_cycles, total_cycles):
    if not total_cycles:
        return None
    return 100.0 * total_full_cycles / total_cycles


def write_standard_workbook(records, output_file, max_occupancy):
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for cache_name in CACHE_LEVELS:
        worksheet = workbook.create_sheet(STREAK_SHEET_NAMES[cache_name])
        headers = ["Trace"]
        for label, _, _ in STREAK_BINS:
            headers.append(f"{label} Count")
        for label, _, _ in STREAK_BINS:
            headers.append(f"{label} Cycle Total")
        headers.append("Total Cycles/KI")
        headers.append("% of Cycle MSHR is Full")
        for label, _, _ in STREAK_BINS:
            headers.append(f"{label} Cycle %")
        worksheet.append(headers)

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for record in sorted(records, key=lambda item: natural_sort_key(item["trace"])):
            streak_lengths = record["streaks"].get(cache_name, [])
            total_full_cycles = sum(streak_lengths)
            counts = streak_bin_counts(streak_lengths)
            cycles = streak_bin_cycles(streak_lengths)
            total_cycles_per_ki = (
                sum(cycles) / CYCLES_PER_KILO_INSTRUCTION_DIVISOR
            )
            full_cycle_percentage = mshr_full_cycle_percentage(
                total_full_cycles, record["total_cycles"]
            )
            percentages = streak_bin_cycle_percentages(cycles, total_full_cycles)
            row = [record["trace"]]
            row.extend(counts)
            row.extend(cycles)
            row.append(total_cycles_per_ki)
            row.append(full_cycle_percentage if full_cycle_percentage is not None else "N/A")
            row.extend(percentages)
            worksheet.append(row)

        total_cycles_per_ki_col = 2 + (2 * len(STREAK_BINS))
        for row in worksheet.iter_rows(min_row=2, min_col=total_cycles_per_ki_col):
            for cell in row:
                cell.number_format = "0.0000"

        worksheet.freeze_panes = "B2"
        worksheet.column_dimensions["A"].width = 45
        for col_idx in range(2, (len(STREAK_BINS) * 3) + 3):
            worksheet.column_dimensions[worksheet.cell(row=1, column=col_idx).column_letter].width = 12

    workbook.save(output_file)


def safe_sheet_name(configuration, used_names):
    # Excel sheet names cannot contain path separators and are limited to 31 characters.
    name = re.sub(r"[\\/*?:\[\]]", "_", configuration).strip("_") or "results"
    name = name[:31]
    candidate = name
    suffix = 2
    while candidate.lower() in used_names:
        suffix_text = f"_{suffix}"
        candidate = f"{name[:31 - len(suffix_text)]}{suffix_text}"
        suffix += 1
    used_names.add(candidate.lower())
    return candidate


def append_cache_section(
    worksheet, cache_name, records, max_occupancy, title_row=None
):
    if title_row is None:
        title_row = worksheet.max_row + 1
    worksheet.cell(title_row, 1, cache_name)
    worksheet.cell(title_row, 1).font = Font(bold=True)

    streak_header_row = worksheet.max_row + 1
    streak_headers = ["Trace"]
    streak_headers.extend(f"{label} Count" for label, _, _ in STREAK_BINS)
    streak_headers.extend(f"{label} Cycle Total" for label, _, _ in STREAK_BINS)
    streak_headers.append("Total Cycles/KI")
    streak_headers.append("% of Cycle MSHR is Full")
    streak_headers.extend(f"{label} Cycle %" for label, _, _ in STREAK_BINS)
    worksheet.append(streak_headers)
    for cell in worksheet[streak_header_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    total_cycles_per_ki_col = 2 + (2 * len(STREAK_BINS))
    for record in sorted(records, key=lambda item: natural_sort_key(item["trace_name"])):
        streak_lengths = record["streaks"].get(cache_name, [])
        if cache_name not in record["streaks"]:
            continue
        counts = streak_bin_counts(streak_lengths)
        cycles = streak_bin_cycles(streak_lengths)
        total_full_cycles = sum(streak_lengths)
        total_cycles_per_ki = (
            sum(cycles) / CYCLES_PER_KILO_INSTRUCTION_DIVISOR
        )
        full_cycle_percentage = mshr_full_cycle_percentage(
            total_full_cycles, record["total_cycles"]
        )
        percentages = streak_bin_cycle_percentages(
            cycles, total_full_cycles
        )
        row = (
            [record["trace_name"]]
            + counts
            + cycles
            + [total_cycles_per_ki]
            + [full_cycle_percentage if full_cycle_percentage is not None else "N/A"]
            + percentages
        )
        worksheet.append(row)
        for cell in worksheet[worksheet.max_row][total_cycles_per_ki_col - 1:]:
            cell.number_format = "0.0000"


def write_grouped_workbook(records, output_file, max_occupancy):
    grouped_records = defaultdict(list)
    for record in records:
        grouped_records[record["configuration"]].append(record)

    workbook = Workbook()
    workbook.remove(workbook.active)
    used_names = set()

    for configuration in sorted(grouped_records, key=natural_sort_key):
        worksheet = workbook.create_sheet(safe_sheet_name(configuration, used_names))
        worksheet.cell(1, 1, configuration.replace(os.sep, "/"))
        worksheet.cell(1, 1).font = Font(bold=True)

        for index, cache_name in enumerate(CACHE_LEVELS):
            title_row = worksheet.max_row + (4 if index else 1)
            append_cache_section(
                worksheet,
                cache_name,
                grouped_records[configuration],
                max_occupancy,
                title_row,
            )

        worksheet.freeze_panes = "B3"
        worksheet.column_dimensions["A"].width = 48
        max_columns = (len(STREAK_BINS) * 3) + 2
        for col_idx in range(2, max_columns + 1):
            column_letter = worksheet.cell(row=1, column=col_idx).column_letter
            worksheet.column_dimensions[column_letter].width = 14

    workbook.save(output_file)


def write_workbook(records, output_file, max_occupancy):
    configurations = {record["configuration"] for record in records}
    if len(configurations) == 1:
        write_standard_workbook(records, output_file, max_occupancy)
    else:
        write_grouped_workbook(records, output_file, max_occupancy)


def main():
    parser = argparse.ArgumentParser(
        description="Extract MSHR occupancy distribution buckets from ChampSim output files."
    )
    parser.add_argument(
        "results_dir",
        nargs="?",
        default=DEFAULT_RESULTS_DIR,
        help=f"Directory containing ChampSim output files. Default: {DEFAULT_RESULTS_DIR}",
    )
    parser.add_argument(
        "-o",
        "--output",
        default=DEFAULT_OUTPUT_FILE,
        help=f"Excel output file. Default: {DEFAULT_OUTPUT_FILE}",
    )
    parser.add_argument(
        "--max-occupancy",
        type=int,
        default=16,
        help="Highest MSHR occupancy column to write. Default: 16",
    )
    args = parser.parse_args()

    if not os.path.isdir(args.results_dir):
        raise SystemExit(f"Results directory not found: {args.results_dir}")

    records = collect_records(args.results_dir)
    if not records:
        raise SystemExit(
            "No parseable result files found under: "
            f"{args.results_dir}\n"
            "Expected lines such as 'L1D MSHR FULL STREAK LENGTHS:'."
        )

    write_workbook(records, args.output, args.max_occupancy)
    print(f"Wrote {len(records)} traces to {args.output}")


if __name__ == "__main__":
    main()
