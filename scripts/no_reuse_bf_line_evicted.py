#!/usr/bin/env python3
import argparse
import os
import re

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
except ImportError:
    raise SystemExit("The 'openpyxl' library is required. Install it with: pip install openpyxl")


DEFAULT_RESULTS_DIR = "../results/ai_ml/Reuse_Count_all_packets"
DEFAULT_OUTPUT_FILE = "line_reuse_count.xlsx"
CACHE_LEVELS = ("L1D", "L2C", "LLC")
SHEET_NAMES = {"L1D": "L1D", "L2C": "L2", "LLC": "LLC"}
DEFAULT_WAYS = {"L1D": 4, "L2C": 6, "LLC": 3}


def natural_sort_key(value):
    return [int(text) if text.isdigit() else text.lower()
            for text in re.split(r"([0-9]+)", value)]


def load_cache_ways(cache_header):
    ways = DEFAULT_WAYS.copy()
    if not os.path.exists(cache_header):
        return ways

    with open(cache_header, "r", errors="ignore") as file:
        content = file.read()

    for cache_name in CACHE_LEVELS:
        match = re.search(rf"#define\s+{cache_name}_WAY\s+(\d+)", content)
        if match:
            ways[cache_name] = int(match.group(1))

    return ways


def parse_reuse_counts(histogram_text):
    counts = {}
    for reuse_count, num_lines in re.findall(r"(\d+)\s*:\s*(\d+)", histogram_text):
        counts[int(reuse_count)] = int(num_lines)
    return counts


def parse_file(filepath):
    with open(filepath, "r", errors="ignore") as file:
        content = file.read()

    reuse_counts = {}
    for cache_name in CACHE_LEVELS:
        match = re.search(
            rf"{cache_name}\s*LINE REUSE COUNT\s*:\s*([^\n\r]*)",
            content,
        )
        if match:
            reuse_counts[cache_name] = parse_reuse_counts(match.group(1))

    return reuse_counts


def collect_records(results_dir):
    records = []
    for root, _, files in os.walk(results_dir):
        for filename in sorted(files, key=natural_sort_key):
            filepath = os.path.join(root, filename)
            if not os.path.isfile(filepath):
                continue

            reuse_counts = parse_file(filepath)
            if not reuse_counts:
                continue

            records.append({
                "trace": os.path.basename(filepath),
                "reuse_counts": reuse_counts,
            })

    return records


def build_reuse_row(trace, counts, max_reuse_column):
    count_values = [counts.get(reuse_count, 0) for reuse_count in range(max_reuse_column + 1)]
    overflow = sum(num_lines for reuse_count, num_lines in counts.items()
                   if reuse_count > max_reuse_column)
    total_lines = sum(count_values) + overflow

    if total_lines:
        percentage_values = [(100.0 * count / total_lines) for count in count_values]
        overflow_percentage = 100.0 * overflow / total_lines
    else:
        percentage_values = [0.0] * len(count_values)
        overflow_percentage = 0.0

    row = [trace]
    row.extend(count_values)
    row.append(overflow)
    row.extend(percentage_values)
    row.append(overflow_percentage)
    return row


def write_workbook(records, output_file, cache_ways):
    workbook = Workbook()
    workbook.remove(workbook.active)

    for cache_name in CACHE_LEVELS:
        max_reuse_column = cache_ways[cache_name] 
        worksheet = workbook.create_sheet(SHEET_NAMES[cache_name])
        headers = ["Trace"] + [str(value) for value in range(max_reuse_column + 1)]
        headers.append(f">{max_reuse_column}")
        headers.extend([f"{value} %" for value in range(max_reuse_column + 1)])
        headers.append(f">{max_reuse_column} %")
        worksheet.append(headers)

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for record in sorted(records, key=lambda item: natural_sort_key(item["trace"])):
            counts = record["reuse_counts"].get(cache_name)
            if counts is None:
                continue
            worksheet.append(build_reuse_row(record["trace"], counts, max_reuse_column))

        first_percentage_col = max_reuse_column 
        for row in worksheet.iter_rows(min_row=2, min_col=first_percentage_col):
            for cell in row:
                cell.number_format = "0.0000"

        worksheet.freeze_panes = "B2"
        worksheet.column_dimensions["A"].width = 45
        for col_idx in range(2, (2 * (max_reuse_column + 2)) + 1):
            worksheet.column_dimensions[worksheet.cell(row=1, column=col_idx).column_letter].width = 12

    workbook.save(output_file)


def main():
    parser = argparse.ArgumentParser(
        description="Extract LINE REUSE COUNT histograms from ChampSim output files."
    )
    parser.add_argument(
        "results_dir",
        nargs="?",
        default=DEFAULT_RESULTS_DIR,
        help=f"Directory containing ChampSim output files. Default: {DEFAULT_RESULTS_DIR}",
    )
    parser.add_argument(
        "-o",
        "--output",
        default=DEFAULT_OUTPUT_FILE,
        help=f"Excel output file. Default: {DEFAULT_OUTPUT_FILE}",
    )
    parser.add_argument(
        "--cache-header",
        default="inc/cache.h",
        help="Path to cache.h for reading *_WAY constants. Default: inc/cache.h",
    )
    args = parser.parse_args()

    if not os.path.isdir(args.results_dir):
        raise SystemExit(f"Results directory not found: {args.results_dir}")

    cache_ways = load_cache_ways(args.cache_header)
    records = collect_records(args.results_dir)
    write_workbook(records, args.output, cache_ways)
    print(f"Wrote {len(records)} traces to {args.output}")


if __name__ == "__main__":
    main()
