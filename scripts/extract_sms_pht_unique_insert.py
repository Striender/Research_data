#!/usr/bin/env python3
"""Extract ``sms.pht.unique_insert`` from ChampSim result traces."""

from __future__ import annotations

import argparse
import os
import re
import shutil
import tempfile
from collections import defaultdict
from pathlib import Path


METRIC = "sms.pht.unique_insert"
METRIC_PATTERN = re.compile(r"^\s*sms\.pht\.unique_insert\s+([+-]?\d+)\s*$", re.MULTILINE)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_RESULTS_DIR = os.path.normpath(
    os.path.join(SCRIPT_DIR, "..", "results", "SMS_region", "4k")
)
DEFAULT_OUTPUT_FILE = os.path.normpath(
    os.path.join(SCRIPT_DIR, "..", "Excel_Output", "sms", "Unique_PHT_entries.xlsx")
)


def natural_sort_key(value: str) -> list[object]:
    """Sort names such as exp2 before exp10."""
    return [
        int(part) if part.isdigit() else part.casefold()
        for part in re.split(r"(\d+)", value)
    ]


def extract_unique_insert(trace_path: Path) -> int | None:
    """Return the last unique-insert count printed in one trace file."""
    try:
        content = trace_path.read_text(errors="ignore")
    except OSError as exc:
        print(f"Warning: could not read {trace_path}: {exc}")
        return None

    matches = METRIC_PATTERN.findall(content)
    return int(matches[-1]) if matches else None


def collect_results(results_dir: Path) -> tuple[dict[str, list[dict[str, object]]], int]:
    """Collect metrics using the same directory grouping as extract_IPC.py."""
    grouped: dict[str, list[dict[str, object]]] = defaultdict(list)
    scanned = 0

    for root, _dirs, files in os.walk(results_dir):
        if not files:
            continue
        relative_path = os.path.relpath(root, results_dir)
        path_parts = relative_path.split(os.sep)
        group_key = None
        experiment = None

        if len(path_parts) == 4:
            cache, size, prefetcher, experiment = path_parts
            group_key = f"{cache}_{size}_{prefetcher}"
        elif len(path_parts) == 3:
            cache_level, prefetcher, experiment = path_parts
            group_key = f"{cache_level}_{prefetcher}"
        elif len(path_parts) == 2:
            cache_level, experiment = path_parts
            if cache_level == "no_pref":
                group_key = cache_level

        if not group_key or not experiment:
            continue

        for filename in sorted(files, key=natural_sort_key):
            trace_path = Path(root) / filename
            scanned += 1
            value = extract_unique_insert(trace_path)
            if value is not None:
                grouped[group_key].append(
                    {
                        "Experiment": experiment,
                        "Trace File": filename,
                        METRIC: value,
                    }
                )

    for records in grouped.values():
        records.sort(
            key=lambda record: (
                natural_sort_key(str(record["Experiment"])),
                natural_sort_key(str(record["Trace File"])),
            )
        )
    return dict(grouped), scanned


def apply_border_to_range(worksheet, row_range, col_range, border) -> None:
    for row in worksheet.iter_rows(
        min_row=row_range[0], max_row=row_range[1],
        min_col=col_range[0], max_col=col_range[1],
    ):
        for cell in row:
            cell.border = border


def write_workbook(
    grouped: dict[str, list[dict[str, object]]], output_path: Path
) -> None:
    """Write per-prefetcher sheets and a summary in extract_IPC.py format."""
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.cell import MergedCell
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError(
            "The 'openpyxl' package is required. Install it with: pip install openpyxl"
        ) from exc

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.active.title = "Placeholder"
    generated_names = set(grouped) | {"Unique_Insert_Summary", "Placeholder"}
    custom_sheets = []
    if output_path.exists():
        old_workbook = load_workbook(output_path)
        for sheet_name in old_workbook.sheetnames:
            if not sheet_name.startswith("raw_") and sheet_name not in generated_names:
                old_sheet = old_workbook[sheet_name]
                custom_sheets.append(
                    (sheet_name, [[cell.value for cell in row] for row in old_sheet.iter_rows()])
                )

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    main_header_fill = PatternFill("solid", fgColor="FFDAB9")
    data_header_fill = PatternFill("solid", fgColor="A9A9A9")
    sub_header_fill = PatternFill("solid", fgColor="ADD8E6")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    headers = ["Trace File", METRIC]

    for group_key in sorted(grouped, key=natural_sort_key):
        worksheet = workbook.create_sheet(group_key[:31])
        if group_key == "no_pref":
            title = "Baseline (No Prefetcher)"
        else:
            parts = group_key.split("_")
            cache_level = parts[1].upper() if len(parts) > 1 else ""
            prefetcher = "_".join(parts[2:]).capitalize() if len(parts) > 2 else parts[0].capitalize()
            title = f"Data Prefetcher: {prefetcher} at {cache_level}"

        worksheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=2)
        title_cell = worksheet.cell(1, 1, title)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = center
        title_cell.fill = main_header_fill
        apply_border_to_range(worksheet, (1, 2), (1, 2), thin_border)

        for column, header in enumerate(headers, 1):
            cell = worksheet.cell(3, column, header)
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.fill = data_header_fill
            cell.border = thin_border
            cell.alignment = center
        worksheet.row_dimensions[3].height = 30
        worksheet.freeze_panes = "A4"

        current_row = 3
        experiments = sorted(
            {str(record["Experiment"]) for record in grouped[group_key]},
            key=natural_sort_key,
        )
        for experiment in experiments:
            records = [r for r in grouped[group_key] if r["Experiment"] == experiment]
            if current_row > 3:
                current_row += 1
            try:
                exp_parts = experiment.split("_")
                exp_number = "".join(filter(str.isdigit, exp_parts[0]))
                exp_title = (
                    f"Experiment {exp_number}: Replacement Policy {exp_parts[1].upper()} "
                    f"at L2 and {exp_parts[2].upper()} at LLC"
                )
            except IndexError:
                exp_title = experiment.replace("_", " ").title()

            worksheet.merge_cells(
                start_row=current_row + 1, start_column=1,
                end_row=current_row + 1, end_column=2,
            )
            cell = worksheet.cell(current_row + 1, 1, exp_title)
            cell.font = Font(bold=True, size=12)
            cell.alignment = center
            cell.fill = sub_header_fill
            apply_border_to_range(
                worksheet, (current_row + 1, current_row + 1), (1, 2), thin_border
            )
            worksheet.row_dimensions[current_row + 1].height = 30

            for row_number, record in enumerate(records, current_row + 2):
                trace_cell = worksheet.cell(row_number, 1, record["Trace File"])
                trace_cell.font = Font(bold=True)
                trace_cell.alignment = left
                worksheet.cell(row_number, 2, record[METRIC]).alignment = right
            current_row += 1 + len(records)

        for column in range(1, 3):
            letter = get_column_letter(column)
            values = (
                str(cell.value) for cell in worksheet[letter]
                if not isinstance(cell, MergedCell) and cell.value is not None
            )
            worksheet.column_dimensions[letter].width = max(map(len, values), default=0) + 2

    summary = workbook.create_sheet("Unique_Insert_Summary", 0)
    summary_headers = ["Experiment", "Trace File"] + sorted(grouped, key=natural_sort_key)
    for column, header in enumerate(summary_headers, 1):
        cell = summary.cell(1, column, header)
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = data_header_fill
        cell.border = thin_border
        cell.alignment = center

    summary_values: dict[tuple[str, str], dict[str, object]] = defaultdict(dict)
    for group_key, records in grouped.items():
        for record in records:
            key = (str(record["Experiment"]), str(record["Trace File"]))
            summary_values[key][group_key] = record[METRIC]
    for row_number, key in enumerate(
        sorted(summary_values, key=lambda x: (natural_sort_key(x[0]), natural_sort_key(x[1]))), 2
    ):
        values = [key[0], key[1]] + [summary_values[key].get(g) for g in summary_headers[2:]]
        for column, value in enumerate(values, 1):
            cell = summary.cell(row_number, column, value)
            cell.border = thin_border
            cell.alignment = left if column <= 2 else right
    summary.freeze_panes = "C2"
    summary.auto_filter.ref = summary.dimensions
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 55
    for column in range(3, len(summary_headers) + 1):
        summary.column_dimensions[get_column_letter(column)].width = 24

    for sheet_name, rows in custom_sheets:
        sheet = workbook.create_sheet(sheet_name)
        for row in rows:
            sheet.append(row)
    workbook.remove(workbook["Placeholder"])

    temporary = Path(tempfile.gettempdir()) / f"{os.getpid()}_{output_path.name}"
    workbook.save(temporary)
    shutil.move(temporary, output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=f"Extract {METRIC} values from every trace below a results directory."
    )
    parser.add_argument(
        "results_dir",
        nargs="?",
        type=Path,
        default=Path(DEFAULT_RESULTS_DIR),
        help=f"directory to scan recursively (default: {DEFAULT_RESULTS_DIR})",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=Path(DEFAULT_OUTPUT_FILE),
        help=f"output .xlsx file (default: {DEFAULT_OUTPUT_FILE})",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    results_dir = args.results_dir.resolve()
    output_path = args.output.resolve()

    if not results_dir.is_dir():
        print(f"Error: results directory not found: {results_dir}")
        return 1
    if output_path.suffix.lower() != ".xlsx":
        print("Error: output filename must end in .xlsx")
        return 1

    grouped, scanned = collect_results(results_dir)
    extracted = sum(len(records) for records in grouped.values())
    if not grouped:
        print(f"No {METRIC} values found in {results_dir}")
        return 1
    try:
        write_workbook(grouped, output_path)
    except RuntimeError as exc:
        print(f"Error: {exc}")
        return 1
    print(
        f"Extracted {extracted} {METRIC} values from {scanned} files "
        f"into {len(grouped)} group sheet(s).\n"
        f"Workbook written to: {output_path}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
