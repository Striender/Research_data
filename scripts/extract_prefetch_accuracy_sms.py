#!/usr/bin/env python3
"""Extract L1D and L2C prefetch accuracy from SMS ChampSim results."""

import argparse
import os
import re

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
except ImportError:
    raise SystemExit(
        "The 'openpyxl' library is required. Install it with: pip install openpyxl"
    )


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_RESULTS_DIR = os.path.normpath(
    os.path.join(SCRIPT_DIR, "..", "results", "sms_ai_ml", "with-sms")
)
DEFAULT_OUTPUT_FILE = os.path.join(SCRIPT_DIR, "sms_prefetch_accuracy.xlsx")
CACHE_LEVELS = ("L1D", "L2C")


def natural_sort_key(value):
    return [
        int(part) if part.isdigit() else part.lower()
        for part in re.split(r"(\d+)", value)
    ]


def parse_counter(content, cache_name, counter_name):
    values = re.findall(
        rf"^Core_\d+_{cache_name}_prefetch_{counter_name}\s+(\d+)\s*$",
        content,
        flags=re.MULTILINE,
    )
    return sum(map(int, values)) if values else None


def parse_file(filepath):
    with open(filepath, "r", errors="ignore") as result_file:
        content = result_file.read()

    cache_stats = {}
    for cache_name in CACHE_LEVELS:
        issued = parse_counter(content, cache_name, "issued")
        useful = parse_counter(content, cache_name, "useful")
        if issued is None or useful is None:
            continue

        accuracy = useful * 100.0 / issued if issued else None
        cache_stats[cache_name] = {
            "issued": issued,
            "useful": useful,
            "accuracy": accuracy,
        }

    return cache_stats


def collect_records(results_dir):
    records = []
    for root, directories, files in os.walk(results_dir):
        directories.sort(key=natural_sort_key)
        for filename in sorted(files, key=natural_sort_key):
            filepath = os.path.join(root, filename)
            if not os.path.isfile(filepath):
                continue

            cache_stats = parse_file(filepath)
            if not cache_stats:
                continue

            relpath = os.path.relpath(filepath, results_dir)
            records.append({
                "trace": relpath,
                "stats": cache_stats,
            })

    return records


def write_workbook(records, output_file):
    workbook = Workbook()
    workbook.remove(workbook.active)

    for cache_name in CACHE_LEVELS:
        worksheet = workbook.create_sheet(cache_name)
        accuracies = []
        worksheet.append([
            "Trace",
            "Prefetch Useful",
            "Prefetch Issued",
            "Prefetch Accuracy (%)",
        ])

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for record in sorted(records, key=lambda item: natural_sort_key(item["trace"])):
            stats = record["stats"].get(cache_name)
            if stats is None:
                continue

            worksheet.append([
                record["trace"],
                stats["useful"],
                stats["issued"],
                stats["accuracy"] if stats["accuracy"] is not None else "N/A",
            ])
            if stats["accuracy"] is not None:
                worksheet.cell(worksheet.max_row, 4).number_format = "0.0000"
                accuracies.append(stats["accuracy"])

        worksheet.append([])
        worksheet.append([
            "Average Accuracy",
            None,
            None,
            sum(accuracies) / len(accuracies) if accuracies else "N/A",
        ])
        average_row = worksheet.max_row
        worksheet.cell(average_row, 1).font = Font(bold=True)
        worksheet.cell(average_row, 4).font = Font(bold=True)
        if accuracies:
            worksheet.cell(average_row, 4).number_format = "0.0000"

        worksheet.freeze_panes = "B2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.column_dimensions["A"].width = 60
        worksheet.column_dimensions["B"].width = 20
        worksheet.column_dimensions["C"].width = 20
        worksheet.column_dimensions["D"].width = 24

    workbook.save(output_file)


def main():
    parser = argparse.ArgumentParser(
        description="Calculate L1D and L2C prefetch accuracy from ChampSim results."
    )
    parser.add_argument(
        "results_dir",
        nargs="?",
        default=DEFAULT_RESULTS_DIR,
        help=f"Results directory. Default: {DEFAULT_RESULTS_DIR}",
    )
    parser.add_argument(
        "-o",
        "--output",
        default=DEFAULT_OUTPUT_FILE,
        help=f"Excel output file. Default: {DEFAULT_OUTPUT_FILE}",
    )
    args = parser.parse_args()

    if not os.path.isdir(args.results_dir):
        raise SystemExit(f"Results directory not found: {args.results_dir}")

    records = collect_records(args.results_dir)
    if not records:
        raise SystemExit(
            f"No prefetch counters found under: {args.results_dir}\n"
            "Expected keys such as 'Core_0_L1D_prefetch_issued'."
        )

    output_file = os.path.abspath(args.output)
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    write_workbook(records, output_file)
    print(f"Wrote {len(records)} traces to {output_file}")


if __name__ == "__main__":
    main()
